import openpyxl
import requests
from bs4 import BeautifulSoup
from datetime import datetime

def get_longest_shortest_options(keyword):
    url = f"https://www.google.com/search?q={keyword}"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')

    # Extracting the search results
    search_results = soup.find_all('div', class_='BNeawe')

    # Finding the longest and shortest options
    longest_option = min(search_results, key=lambda x: len(x.text)).text
    shortest_option = max(search_results, key=lambda x: len(x.text)).text

    return longest_option, shortest_option

def main():
    # Load the Excel file
    excel_file_path = r'D:\Python\q1.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)

    # Get the current day of the week (Monday, Tuesday, etc.)
    current_day = datetime.now().strftime('%A')

    try:
        # Select the worksheet corresponding to the current day
        worksheet = workbook[current_day]
    except KeyError:
        print(f"Worksheet for {current_day} does not exist in the Excel file.")
        return

    # Iterate through each row in the worksheet
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]

        # Get longest and shortest options
        longest_option, shortest_option = get_longest_shortest_options(keyword)

        # Check if row[1] is a valid integer before updating the Excel file
        try:
            row_index = int(row[1]) if row[1] is not None and row[1].isdigit() else None
        except ValueError:
            print(f"Invalid value in row[1]: {row[1]}")
            continue

        # Check if row_index is not None before updating the Excel file
        if row_index is not None:
            worksheet.cell(row=row_index, column=2, value=longest_option)
            worksheet.cell(row=row_index, column=3, value=shortest_option)

    # Save the updated Excel file
    workbook.save(r'D:\Python\updated_sample_excel_file.xlsx')

if __name__ == "__main__":
    main()
