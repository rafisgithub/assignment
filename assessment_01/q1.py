from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
import time
from datetime import datetime, timedelta

def main():
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://google.com/")

    keywords = get_keywords()
    shortest_list = []
    longest_list = []

    for keyword in keywords:
        search_box = driver.find_element(By.ID, "id")
        search_box.send_keys(keyword)
        time.sleep(1)

        elements = driver.find_elements(By.XPATH, "//span")
        element_texts = [element.text for element in elements if element.text.strip()]
        shortest_list.append(shortest_string(element_texts))
        longest_list.append(longest_string(element_texts))

        search_box.clear()
        time.sleep(1)

    write_to_excel(shortest_list, 3)
    write_to_excel(longest_list, 4)

    driver.quit()

def get_keywords():

    file_path = "D:\\Python\\AssignementOnSelenium\\Excel.xlsx"
    day_of_week_string = day()
    workbook = load_workbook(file_path)
    sheet = workbook[day_of_week_string]
    column_values = [cell.value for row in sheet.iter_rows(min_row=1, max_col=1, min_col=2) for cell in row if cell.value]
    workbook.close()
    return column_values

def shortest_string(strings):
    if not strings:
        return None
    return min(strings, key=len)

def longest_string(strings):
    if not strings:
        return None
    return max(strings, key=len)

def write_to_excel(strings, column_index):
    file_path = "D:\\Python\\AssignementOnSelenium\\Excel.xlsx"
    day_of_week = day()

    workbook = load_workbook(file_path)
    sheet = workbook[day_of_week]

    for i, string in enumerate(strings):
        row_index = i + 2
        sheet.cell(row=row_index, column=column_index).value = string

    workbook.save(file_path)
    print("Excel file has been updated successfully.")

def day():
    current_date = datetime.now()
    day_of_week_string = current_date.strftime("%A")
    return day_of_week_string

if __name__ == "__main__":
    main()
